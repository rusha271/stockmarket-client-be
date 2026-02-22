"""
Cron job: every 5 min during NSE hours (Mon–Fri 9:15–15:30 IST).
1. Evaluate previous run: load run from 5 min ago, fetch actual prices, compare, append to Excel.
2. Run predictions for all Nifty 50, save run for next evaluation.
Saves to D:\\be-stock\\app\\data (or app/data) with date in file names.
"""

from __future__ import annotations

import json
import logging
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from app.nifty50 import NIFTY_50_SYMBOLS
from app.services.next_timeframe import MARKET_OPEN, MARKET_CLOSE

IST = ZoneInfo("Asia/Kolkata")
DATA_DIR = Path(__file__).resolve().parent / "data"
RUNS_DIR = DATA_DIR / "runs"

logger = logging.getLogger(__name__)


def _ensure_dirs():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    RUNS_DIR.mkdir(parents=True, exist_ok=True)


def _run_filename(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%d_%H-%M") + ".json"


def _excel_filename(d: datetime | None) -> str:
    d = d or datetime.now(IST)
    if hasattr(d, "strftime"):
        return f"predictions_{d.strftime('%Y-%m-%d')}.xlsx"
    return f"predictions_{d}.xlsx"


def _get_actual_price(symbol: str) -> float | None:
    """Fetch latest 5m close for symbol from Yahoo (actual at evaluation time). Use same fetch as prediction so data is available."""
    try:
        from app.services.yahoo_ohlc import fetch_ohlc
        candles = fetch_ohlc(symbol, days=5, interval="5m", last_n=10)
        if candles:
            return float(candles[-1]["close"])
    except Exception as e:
        logger.debug("Could not fetch actual for %s: %s", symbol, e)
    return None


def run_predictions_for_all(clf, reg) -> list[dict]:
    """Run prediction for each Nifty 50 symbol; return list of {symbol, current_price, predicted_price}."""
    from app.services.yahoo_ohlc import fetch_ohlc
    from app.services.feature_builder import build_features_from_candles
    from app.services.candlestick_service import predict_candlestick, get_num_features
    from app.services.next_timeframe import is_peak_hours_ist

    results = []
    use_small_bins = not is_peak_hours_ist()
    n = get_num_features(clf) or get_num_features(reg) or 49

    for symbol in NIFTY_50_SYMBOLS:
        try:
            candles = fetch_ohlc(symbol, days=6, interval="5m", last_n=40)
            if len(candles) < 8:
                continue
            features = build_features_from_candles(candles, n)
            current_close = candles[-1]["close"]
            if current_close <= 0:
                continue
            cr = predict_candlestick(features, current_close, clf, reg, use_5min_bins=use_small_bins)
            predicted = round(current_close * (1 + cr.price_change_pct / 100.0), 2)
            results.append({
                "symbol": symbol,
                "current_price": round(current_close, 2),
                "predicted_price": predicted,
            })
        except Exception as e:
            logger.warning("Prediction failed for %s: %s", symbol, e)

    return results


def save_run(prediction_time: datetime, results: list[dict]) -> Path:
    """Save run JSON to data/runs/YYYY-MM-DD_HH-MM.json."""
    _ensure_dirs()
    path = RUNS_DIR / _run_filename(prediction_time)
    payload = {
        "prediction_time": prediction_time.isoformat(),
        "predicted_for_5min_later": (prediction_time + timedelta(minutes=5)).isoformat(),
        "symbols": results,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)
    logger.info("Saved run %s (%d symbols) to %s", prediction_time.strftime("%H:%M"), len(results), path)
    return path


def load_run(prediction_time: datetime) -> list[dict] | None:
    """Load run JSON; return list of {symbol, current_price, predicted_price} or None."""
    path = RUNS_DIR / _run_filename(prediction_time)
    if not path.exists():
        return None
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    return data.get("symbols")


def evaluate_and_append_to_excel(prediction_time: datetime, date_ist: datetime) -> None:
    """
    Load run from prediction_time, fetch actual for each symbol, compute diffs, append to Excel.
    prediction_time e.g. 13:00, actual is at 13:05.
    """
    symbols_data = load_run(prediction_time)
    if not symbols_data:
        logger.debug("No run file for %s; skip evaluate", prediction_time.strftime("%Y-%m-%d %H:%M"))
        return

    actual_time = prediction_time + timedelta(minutes=5)
    rows = []
    for row in symbols_data:
        symbol = row["symbol"]
        current_at_pred = row["current_price"]
        predicted = row["predicted_price"]
        actual = _get_actual_price(symbol)
        if actual is None:
            continue
        diff_predicted_actual = round(actual - predicted, 2)
        diff_actual_previous = round(actual - current_at_pred, 2)
        pct_pred = round((diff_predicted_actual / predicted * 100), 4) if predicted else None
        pct_prev = round((diff_actual_previous / current_at_pred * 100), 4) if current_at_pred else None
        predicted_direction = "up" if predicted > current_at_pred else ("down" if predicted < current_at_pred else "neutral")
        actual_direction = "up" if actual > current_at_pred else ("down" if actual < current_at_pred else "neutral")
        rows.append({
            "prediction_time": prediction_time.strftime("%H:%M"),
            "actual_time": actual_time.strftime("%H:%M"),
            "symbol": symbol,
            "current_at_prediction": current_at_pred,
            "predicted_price": predicted,
            "predicted_direction": predicted_direction,
            "actual_price": round(actual, 2),
            "actual_direction": actual_direction,
            "diff_predicted_actual": diff_predicted_actual,
            "diff_actual_previous": diff_actual_previous,
            "pct_diff_predicted": pct_pred,
            "pct_diff_previous": pct_prev,
        })

    if not rows:
        logger.warning("No rows to append for run %s (all actuals failed for %d symbols)", prediction_time.strftime("%H:%M"), len(symbols_data))
        return

    _ensure_dirs()
    excel_path = (DATA_DIR / _excel_filename(date_ist)).resolve()
    try:
        _append_rows_to_excel(excel_path, rows)
        logger.info("Appended %d rows to %s (run %s)", len(rows), excel_path, prediction_time.strftime("%H:%M"))
    except Exception as e:
        logger.exception("Failed to write Excel %s: %s", excel_path, e)


def _append_rows_to_excel(path: Path, rows: list[dict]) -> None:
    """Create or append to Excel; includes predicted_direction and actual_direction (up/down/neutral)."""
    import openpyxl

    headers = [
        "prediction_time", "actual_time", "symbol", "current_at_prediction", "predicted_price",
        "predicted_direction", "actual_price", "actual_direction",
        "diff_predicted_actual", "diff_actual_previous", "pct_diff_predicted", "pct_diff_previous",
    ]

    path = path.resolve()
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        wb = openpyxl.load_workbook(path, read_only=False)
        if "Predictions" in wb.sheetnames:
            ws = wb["Predictions"]
        else:
            ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Predictions"
        ws.append(headers)

    for r in rows:
        ws.append([r.get(h) for h in headers])

    wb.save(str(path))


def run_five_min_job(clf, reg) -> None:
    """
    Called every 5 min during market hours.
    1. Evaluate previous run (from 5 min ago): fetch actuals, append to Excel.
    2. Run predictions for all Nifty 50, save run.
    """
    now = datetime.now(IST)
    date_ist = now.date()
    if now.weekday() >= 5:
        return
    t = now.time()
    if t < MARKET_OPEN or t >= MARKET_CLOSE:
        return

    # Slot at 5-min boundary (9:15, 9:20, ...)
    minute = (now.minute // 5) * 5
    current_slot = now.replace(second=0, microsecond=0, minute=minute)
    previous_slot = current_slot - timedelta(minutes=5)

    # 1) Evaluate previous run
    evaluate_and_append_to_excel(previous_slot, date_ist)

    # 2) Run predictions for next 5 min, save run
    results = run_predictions_for_all(clf, reg)
    if results:
        save_run(current_slot, results)
